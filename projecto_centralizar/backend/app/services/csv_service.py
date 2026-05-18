import csv
import io
from openpyxl import load_workbook
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.models.contact import Contact
from app.schemas.contact import ContactCreate, ContactFilterParams
from app.services import contact_service

from app.models.empresa import Empresa
from app.schemas.empresa import EmpresaCreate

from app.core.view_fields.contact_view_fields import CONTACT_VIEW_FIELDS
from app.core.view_fields.empresa_view_fields import EMPRESA_VIEW_FIELDS
from app.domain.relations import M2M_FIELD_MAP, EMPRESA_M2M_FIELD_MAP


# Domain-specific M2M field maps
CONTACT_M2M_FIELDS = M2M_FIELD_MAP
EMPRESA_M2M_FIELDS = EMPRESA_M2M_FIELD_MAP

# Export-friendly column names and name attributes for M2M relations.
# Keys: internal M2M key -> (csv_column_name, attribute_to_read_from_model)
M2M_EXPORT_META = {
    "campaign_ids": ("campaigns", "nombre"),
    "sector_ids":   ("sectors",   "name"),
    "vertical_ids": ("verticals", "name"),
    "product_ids":  ("products",  "name"),
}

def _m2m_export_columns(field_map: dict) -> list[str]:
    """Return CSV column names for the given M2M field map."""
    return [M2M_EXPORT_META[k][0] for k in field_map]

HEADER_MAP = {
    "cargo": "Cargo",
    "first_name": "Nombre",
    "last_name": "Apellidos",
    "email": "Email",
    "phone": "Telefono",
    "linkedin": "LinkedIn",
    "empresa_nombre": "Empresa_Nombre",
    "empresa_cif": "Empresa_CIF",
    "empresa_web": "Empresa_Web",
    "empresa_email": "Empresa_Email",
    "empresa_phone": "Empresa_Telefono",
    "campaigns": "Campañas",
    "categoria": "Categoría",
    "sectors": "Sectores",
    "verticals": "Verticales",
    "products": "Productos",
    # Empresa native fields
    "nombre": "Nombre",
    "web": "Web",
    "numero_empleados": "Num_Empleados",
    "facturacion": "Facturacion",
    "cnae": "CNAE",
    "facebook": "Facebook",
    "web_competidor_1": "Web_Competidor_1",
    "web_competidor_2": "Web_Competidor_2",
    "web_competidor_3": "Web_Competidor_3",
    "provincia": "Provincia",
    "pais": "Pais",
}

def _translate_row(row: dict[str, Any]) -> dict[str, Any]:
    return {HEADER_MAP.get(k, k): v for k, v in row.items()}

CSV_FIELDS = [
    "cargo",
    # Contact native fields
    *CONTACT_VIEW_FIELDS,
    # Categoría de Cargo (derived from cargo relationship)
    "categoria",
    # Empresa explicit fields
    "empresa_nombre",
    "empresa_cif",
    "empresa_web",
    "empresa_email",
    "empresa_phone",
    # M2M Relationships (exported as names)
    *_m2m_export_columns(CONTACT_M2M_FIELDS),
    *_m2m_export_columns(EMPRESA_M2M_FIELDS),
]

EMPRESA_CSV_FIELDS = EMPRESA_VIEW_FIELDS + _m2m_export_columns(EMPRESA_M2M_FIELDS)

def _contact_to_row(contact: Contact) -> dict[str, Any]:
    row = {}

    # Core
    row["cargo"] = contact.cargo.name if contact.cargo else ""
    row["categoria"] = contact.cargo.categoria.name if contact.cargo and contact.cargo.categoria else ""

    # Contact native fields
    for field in CONTACT_VIEW_FIELDS:
        row[field] = getattr(contact, field, "") or ""

    # Empresa explicit fields (Decoupled access)
    empresa = getattr(contact, "empresa_rel", None)
    row["empresa_nombre"] = empresa.nombre if empresa else ""
    row["empresa_cif"] = empresa.cif if empresa else ""
    row["empresa_web"] = empresa.web if empresa else ""
    row["empresa_email"] = empresa.email if empresa else ""
    row["empresa_phone"] = empresa.phone if empresa else ""

    # Contact M2M (export names, not IDs)
    for key, config in CONTACT_M2M_FIELDS.items():
        col_name, name_attr = M2M_EXPORT_META[key]
        rel_list = getattr(contact, config["relation_name"], [])
        row[col_name] = ",".join(getattr(item, name_attr, "") for item in rel_list)

    # Empresa M2M (export names, not IDs)
    if empresa:
        for key, config in EMPRESA_M2M_FIELDS.items():
            col_name, name_attr = M2M_EXPORT_META[key]
            rel_list = getattr(empresa, config["relation_name"], [])
            row[col_name] = ",".join(getattr(item, name_attr, "") for item in rel_list)
    else:
        for key in EMPRESA_M2M_FIELDS.keys():
            col_name = M2M_EXPORT_META[key][0]
            row[col_name] = ""

    return row


def _empresa_to_row(empresa: Empresa) -> dict[str, Any]:
    row = {}

    for field in EMPRESA_VIEW_FIELDS:
        # Skip FK ID fields — we export human-readable names instead
        if field in ("pais_id", "provincia_id"):
            continue
        row[field] = getattr(empresa, field, "") or ""

    # Export human-readable location names from relationships
    row["pais"] = empresa.pais_rel.name if empresa.pais_rel else ""
    row["provincia"] = empresa.provincia_rel.name if empresa.provincia_rel else ""

    for key, config in EMPRESA_M2M_FIELDS.items():
        col_name, name_attr = M2M_EXPORT_META[key]
        rel_list = getattr(empresa, config["relation_name"], [])
        row[col_name] = ",".join(getattr(item, name_attr, "") for item in rel_list)

    return row


async def export_contacts_csv(items: list[Contact]) -> str:
    """Return CSV string for a list of contacts."""
    output = io.StringIO()
    
    translated_fieldnames = [HEADER_MAP.get(f, f) for f in CSV_FIELDS]
    writer = csv.DictWriter(output, fieldnames=translated_fieldnames, extrasaction="ignore")
    writer.writeheader()
    
    for contact in items:
        row = _contact_to_row(contact)
        writer.writerow(_translate_row(row))
    return output.getvalue()

async def export_csv(session: AsyncSession, filters: ContactFilterParams) -> str:
    """Return CSV string for all contacts matching filters."""
    items = await contact_service.list_contacts_unpaginated(session, filters)
    return await export_contacts_csv(items)

def parse_csv(content: bytes) -> list[dict]:
    """
    Decodes CSV bytes and returns a list of dictionaries with stripped keys and values.
    Empty strings are converted to None.
    """
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []

    for row in reader:
        row = {k.strip(): (v.strip() if v else None) for k, v in row.items()}
        rows.append(row)

    return rows

def parse_xlsx(content: bytes) -> list[dict]:
    """
    Parses XLSX bytes and returns a list of dictionaries.
    Headers are normalized (stripped and lowercased).
    Data values are stripped if they are strings.
    """
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    # Safe header extraction with normalization
    headers = [
        (str(cell.value).strip().lower() if cell.value is not None else None)
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        clean_row = {}
        for k, v in zip(headers, row):
            if not k:
                continue
            if isinstance(v, str):
                v = v.strip()
            clean_row[k] = v
        rows.append(clean_row)

    return rows

def parse_file(content: bytes, filename: str) -> list[dict]:
    """Unified parser that handles CSV and XLSX based on filename extension."""
    filename = filename.lower()
    if filename.endswith(".csv"):
        return parse_csv(content)
    if filename.endswith(".xlsx"):
        return parse_xlsx(content)
    raise ValueError("Unsupported file format. Only .csv and .xlsx are supported.")

async def export_empresas_csv(session: AsyncSession, items: list[Empresa]) -> str:
    """Return CSV string for a list of enterprises."""
    output = io.StringIO()
    
    translated_fieldnames = [HEADER_MAP.get(f, f) for f in EMPRESA_CSV_FIELDS]
    writer = csv.DictWriter(output, fieldnames=translated_fieldnames, extrasaction="ignore")
    writer.writeheader()
    
    for empresa in items:
        row = _empresa_to_row(empresa)
        writer.writerow(_translate_row(row))
    return output.getvalue()
